from typing import List, Optional
from uuid import UUID
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query, Response
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.security import get_current_admin_user
from app.models import User
from app.schemas.student import (
    StudentCreate, StudentUpdate, StudentResponse, StudentListResponse,
    CSVUploadPreview, CSVUploadResult, PasswordResetResponse,
    ClassCreate, ClassUpdate, ClassResponse, ClassListResponse
)
from app.schemas.analytics import (
    AdminDashboardOverview, StudentAnalytics, ClassAnalytics, TestAnalytics
)
from app.schemas.report import ReportGenerationRequest, ReportGenerationResponse
from app.services.student_service import StudentService
from app.services.csv_processor import CSVProcessorService
from app.services.analytics_service import AnalyticsService
from app.services.report_service import ReportService
from app.services.email_service import EmailService


router = APIRouter(prefix="/admin", tags=["admin"])
student_service = StudentService()


# Student Management Endpoints

@router.post("/students/upload", response_model=CSVUploadPreview)
async def upload_students_csv(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_admin: User = Depends(get_current_admin_user)
):
    """Upload and validate CSV file for bulk student creation."""
    # Validate file structure
    await CSVProcessorService.validate_csv_structure(file)

    # Parse CSV data
    students_data = await CSVProcessorService.parse_csv_file(file)

    # Validate student data
    valid_records, errors = await CSVProcessorService.validate_student_data(
        students_data, db
    )

    # Transform valid records to match frontend expectations
    # Map CSV columns to StudentCreate schema
    transformed_records = []
    for record in valid_records:
        # Combine first name and surname for full name
        full_name = f"{record.get('First Name', '')} {record.get('Surname', '')}".strip()

        transformed = {
            'email': record.get('Email Address', ''),
            'full_name': full_name,
            'first_name': record.get('First Name', ''),
            'surname': record.get('Surname', ''),
            'student_id': record.get('Student ID', ''),
            'class_name': record.get('Class ID', ''),
            'year_group': int(record.get('Year Group', 0))
        }
        transformed_records.append(transformed)

    return CSVUploadPreview(
        data=transformed_records,
        errors=errors,
        total_rows=len(students_data),
        valid_rows=len(valid_records),
        invalid_rows=len(errors)
    )


@router.post("/students/bulk-create", response_model=CSVUploadResult)
async def bulk_create_students(
    students: List[StudentCreate],
    db: AsyncSession = Depends(get_db),
    current_admin: User = Depends(get_current_admin_user)
):
    """Create multiple students from validated data."""
    result = await student_service.create_bulk_students(db, students)

    return CSVUploadResult(
        file_name="bulk_upload",
        total_records=result["total"],
        successful_records=len(result["successful"]),
        failed_records=len(result["failed"]),
        errors=result["failed"]
    )


@router.get("/students", response_model=StudentListResponse)
async def list_students(
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=1000),
    search: Optional[str] = Query(None),
    class_id: Optional[UUID] = Query(None),
    year_group: Optional[int] = Query(None, ge=1, le=13),
    status: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_admin: User = Depends(get_current_admin_user)
):
    """Get paginated list of students with optional filters."""
    result = await student_service.get_students(
        db=db,
        page=page,
        limit=limit,
        search=search,
        class_id=class_id,
        year_group=year_group,
        status=status
    )

    # Transform Student objects to StudentResponse objects
    student_responses = []
    for student in result["students"]:
        student_response = StudentResponse(
            id=student.id,
            user_id=student.user_id,
            student_code=student.student_code,
            email=student.user.email,
            full_name=student.user.full_name,
            username=student.user.username,
            class_id=student.class_id,
            year_group=student.year_group,
            enrollment_date=student.enrollment_date,
            status=student.status,
            is_active=student.user.is_active,
            created_at=student.created_at,
            updated_at=student.updated_at,
            class_info=student.class_info
        )
        student_responses.append(student_response)

    return StudentListResponse(
        students=student_responses,
        total=result["total"],
        page=result["page"],
        pages=result["pages"],
        limit=result["limit"]
    )


@router.get("/students/{student_id}", response_model=StudentResponse)
async def get_student(
    student_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_admin: User = Depends(get_current_admin_user)
):
    """Get detailed information about a specific student."""
    student = await student_service.get_student(db, student_id)
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    return StudentResponse(
        id=student.id,
        user_id=student.user_id,
        student_code=student.student_code,
        email=student.user.email,
        full_name=student.user.full_name,
        username=student.user.username,
        class_id=student.class_id,
        year_group=student.year_group,
        enrollment_date=student.enrollment_date,
        status=student.status,
        is_active=student.user.is_active,
        created_at=student.created_at,
        updated_at=student.updated_at,
        class_info=student.class_info
    )


@router.put("/students/{student_id}", response_model=StudentResponse)
async def update_student(
    student_id: UUID,
    student_data: StudentUpdate,
    db: AsyncSession = Depends(get_db),
    current_admin: User = Depends(get_current_admin_user)
):
    """Update student information."""
    student = await student_service.update_student(db, student_id, student_data)
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    return StudentResponse(
        id=student.id,
        user_id=student.user_id,
        student_code=student.student_code,
        email=student.user.email,
        full_name=student.user.full_name,
        username=student.user.username,
        class_id=student.class_id,
        year_group=student.year_group,
        enrollment_date=student.enrollment_date,
        status=student.status,
        is_active=student.user.is_active,
        created_at=student.created_at,
        updated_at=student.updated_at,
        class_info=student.class_info
    )


@router.delete("/students/{student_id}")
async def delete_student(
    student_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_admin: User = Depends(get_current_admin_user)
):
    """Delete a student account."""
    success = await student_service.delete_student(db, student_id)
    if not success:
        raise HTTPException(status_code=404, detail="Student not found")

    return {"message": "Student deleted successfully"}


@router.post("/students/{student_id}/reset-password", response_model=PasswordResetResponse)
async def reset_student_password(
    student_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_admin: User = Depends(get_current_admin_user)
):
    """Reset student password and send email notification."""
    result = await student_service.reset_password(db, student_id)
    if not result["success"]:
        raise HTTPException(status_code=400, detail=result["error"])

    return PasswordResetResponse(
        message=result["message"],
        email_sent=result["email_sent"]
    )


@router.post("/students/{student_id}/generate-report", response_model=ReportGenerationResponse)
async def generate_student_report(
    student_id: UUID,
    report_data: ReportGenerationRequest,
    db: AsyncSession = Depends(get_db),
    current_admin: User = Depends(get_current_admin_user)
):
    """
    Generate a comprehensive PDF report for a student and send it via email.

    This endpoint:
    1. Fetches student data and calculates performance metrics
    2. Generates a PDF report with admin-provided qualitative feedback
    3. Sends the report to the student's email address as an attachment
    """
    try:
        # Get student information
        student = await student_service.get_student(db, student_id)
        if not student:
            raise HTTPException(status_code=404, detail="Student not found")

        # Prepare student dict for services
        student_dict = {
            "email": student.user.email,
            "full_name": student.user.full_name,
            "student_code": student.student_code
        }

        # Initialize services
        report_service = ReportService()
        email_service = EmailService()

        # Generate PDF report
        pdf_buffer = await report_service.generate_student_report(
            db=db,
            student_id=str(student_id),
            report_data={
                "strengths": report_data.strengths,
                "areas_for_improvement": report_data.areas_for_improvement,
                "teacher_comment": report_data.teacher_comment,
                "intervention_recommendation": report_data.intervention_recommendation,
                "next_steps": report_data.next_steps
            }
        )

        # Get performance data for email
        performance_data = await report_service._calculate_performance_metrics(db, str(student_id))
        student_data = await report_service._fetch_student_data(db, str(student_id))

        # Prepare report metadata for email
        from datetime import datetime
        report_metadata = {
            "report_period": datetime.now().strftime("%B %Y"),
            "class_name": student_data.get("class_name", "N/A"),
            "year_group": student_data.get("year_group", "N/A"),
            "overall_average": performance_data.get("overall_average", "N/A"),
            "class_rank": performance_data.get("class_rank", "N/A")
        }

        # Send email with PDF attachment
        email_sent = await email_service.send_student_report(
            student=student_dict,
            report_data=report_metadata,
            pdf_buffer=pdf_buffer
        )

        if not email_sent:
            raise HTTPException(
                status_code=500,
                detail="Report generated but failed to send email"
            )

        return ReportGenerationResponse(
            success=True,
            message="Report generated and sent successfully",
            student_email=student.user.email
        )

    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to generate report: {str(e)}"
        )


@router.get("/students/template/download")
async def download_csv_template(
    current_admin: User = Depends(get_current_admin_user)
):
    """Download CSV template for student upload."""
    template = CSVProcessorService.generate_csv_template()

    return Response(
        content=template.getvalue(),
        media_type="text/csv",
        headers={
            "Content-Disposition": "attachment; filename=student_upload_template.csv"
        }
    )


# Analytics Endpoints

@router.get("/analytics/dashboard", response_model=AdminDashboardOverview)
async def get_admin_dashboard_analytics(
    db: AsyncSession = Depends(get_db),
    current_admin: User = Depends(get_current_admin_user)
):
    """Get comprehensive analytics overview for admin dashboard."""
    try:
        analytics = await AnalyticsService.get_admin_dashboard_overview(db)
        return AdminDashboardOverview(**analytics)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get dashboard analytics: {str(e)}")


@router.get("/analytics/students/{student_id}", response_model=StudentAnalytics)
async def get_student_analytics(
    student_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_admin: User = Depends(get_current_admin_user)
):
    """Get detailed analytics for a specific student."""
    try:
        analytics = await AnalyticsService.get_student_analytics(db, student_id)
        return StudentAnalytics(**analytics)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get student analytics: {str(e)}")


@router.get("/analytics/classes/{class_id}", response_model=ClassAnalytics)
async def get_class_analytics(
    class_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_admin: User = Depends(get_current_admin_user)
):
    """Get comprehensive analytics for a specific class."""
    try:
        analytics = await AnalyticsService.get_class_analytics(db, class_id)
        return ClassAnalytics(**analytics)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get class analytics: {str(e)}")


@router.get("/analytics/tests/{test_id}", response_model=TestAnalytics)
async def get_test_analytics(
    test_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_admin: User = Depends(get_current_admin_user)
):
    """Get detailed analytics for a specific test."""
    try:
        analytics = await AnalyticsService.get_test_analytics(db, test_id)
        return TestAnalytics(**analytics)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get test analytics: {str(e)}")


# Class Management Endpoints

@router.get("/classes", response_model=ClassListResponse)
async def list_classes(
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=100),
    search: Optional[str] = Query(None),
    year_group: Optional[int] = Query(None, ge=1, le=13),
    db: AsyncSession = Depends(get_db),
    current_admin: User = Depends(get_current_admin_user)
):
    """Get paginated list of classes with optional filters."""
    result = await student_service.get_classes(
        db=db,
        page=page,
        limit=limit,
        search=search,
        year_group=year_group
    )

    return ClassListResponse(
        classes=result["classes"],
        total=result["total"],
        page=result["page"],
        pages=result["pages"],
        limit=result["limit"]
    )


@router.get("/classes/{class_id}/students", response_model=StudentListResponse)
async def get_students_by_class(
    class_id: UUID,
    page: int = Query(1, ge=1),
    limit: int = Query(100, ge=1, le=500),
    db: AsyncSession = Depends(get_db),
    current_admin: User = Depends(get_current_admin_user)
):
    """Get all students in a specific class."""
    from sqlalchemy import select, func
    from app.models.student import Student
    from app.models.user import User as UserModel
    from app.models.class_model import Class
    from sqlalchemy.orm import selectinload

    # Verify class exists
    class_result = await db.execute(select(Class).where(Class.id == class_id))
    class_obj = class_result.scalar_one_or_none()

    if not class_obj:
        raise HTTPException(status_code=404, detail="Class not found")

    # Count total students in class
    count_query = select(func.count()).select_from(Student).where(Student.class_id == class_id)
    total_result = await db.execute(count_query)
    total = total_result.scalar()

    # Get students
    offset = (page - 1) * limit
    query = (
        select(Student)
        .options(
            selectinload(Student.user),
            selectinload(Student.class_info)
        )
        .where(Student.class_id == class_id)
        .offset(offset)
        .limit(limit)
        .order_by(UserModel.full_name)
    )

    result = await db.execute(query)
    students = result.scalars().all()

    # Format response
    student_responses = []
    for student in students:
        student_responses.append({
            "id": str(student.id),
            "user_id": str(student.user_id),
            "student_code": student.student_code,
            "email": student.user.email,
            "full_name": student.user.full_name,
            "username": student.user.username,
            "class_id": str(student.class_id) if student.class_id else None,
            "year_group": student.year_group,
            "enrollment_date": student.enrollment_date.isoformat() if student.enrollment_date else None,
            "status": student.status.value,
            "is_active": student.user.is_active,
            "created_at": student.created_at.isoformat() if student.created_at else None,
            "updated_at": student.updated_at.isoformat() if student.updated_at else None,
            "class_info": {
                "id": str(student.class_info.id),
                "name": student.class_info.name,
                "year_group": student.class_info.year_group,
                "academic_year": student.class_info.academic_year
            } if student.class_info else None
        })

    pages = (total + limit - 1) // limit

    return StudentListResponse(
        students=student_responses,
        total=total,
        page=page,
        pages=pages,
        limit=limit
    )


@router.get("/results/by-week", response_model=None)
async def get_results_by_week(
    week: Optional[int] = Query(None, ge=1, le=40, description="Week number (1-40). If not provided, returns current week."),
    class_id: Optional[UUID] = Query(None, description="Filter by class"),
    db: AsyncSession = Depends(get_db),
    current_admin: User = Depends(get_current_admin_user)
):
    """
    Get test results organized by academic week.

    Returns results for a specific week (or current week if not specified),
    organized by student and subject matching the Excel template format.
    """
    from sqlalchemy import select, and_
    from app.models.test import TestResult, Test
    from app.models.student import Student
    from sqlalchemy.orm import selectinload
    from app.models.class_model import Class
    from app.services.academic_calendar_service import calendar_service
    from app.schemas.analytics import (
        WeeklyResultsSummary, StudentWeeklyScores, SubjectScore, AcademicWeekInfo
    )
    from collections import defaultdict

    # Determine which week to show
    target_week = week if week else calendar_service.get_current_week()

    if target_week == 0:
        raise HTTPException(status_code=400, detail="No active academic week")

    # Get week information
    week_info = calendar_service.get_week_info(target_week)
    start_date, end_date = week_info.start_date, week_info.end_date

    # Build query for results in this week
    query = (
        select(TestResult)
        .options(
            selectinload(TestResult.student).selectinload(Student.user),
            selectinload(TestResult.student).selectinload(Student.class_info),
            selectinload(TestResult.test)
        )
        .where(
            and_(
                TestResult.submitted_at >= start_date,
                TestResult.submitted_at <= end_date
            )
        )
    )

    # Apply class filter if provided
    if class_id:
        query = query.join(Student).where(Student.class_id == class_id)

    query = query.order_by(TestResult.submitted_at)

    result = await db.execute(query)
    week_results = result.scalars().all()

    # Organize results by student
    student_data = defaultdict(lambda: {
        "info": {},
        "scores": {}
    })

    subjects = ["English", "VR GL", "NVR", "Maths"]

    for test_result in week_results:
        if not test_result.student or not test_result.test:
            continue

        student = test_result.student
        test = test_result.test
        student_id = str(student.id)

        # Store student info (once per student)
        if not student_data[student_id]["info"]:
            full_name = student.user.full_name if student.user else ""
            name_parts = full_name.strip().split(maxsplit=1)
            first_name = name_parts[0] if len(name_parts) > 0 else ""
            surname = name_parts[1] if len(name_parts) > 1 else ""

            student_data[student_id]["info"] = {
                "student_id": student_id,
                "student_code": student.student_code or "N/A",
                "first_name": first_name,
                "surname": surname,
                "full_name": full_name,
                "class_id": str(student.class_id) if student.class_id else "",
                "class_name": student.class_info.name if student.class_info else "N/A",
                "year_group": student.year_group
            }

        # Map test type to subject name
        test_type = test.type.value if hasattr(test.type, 'value') else str(test.type)
        subject_map = {
            "English": "English",
            "Verbal Reasoning": "VR GL",
            "Non-Verbal Reasoning": "NVR",
            "Mathematics": "Maths"
        }
        subject_name = subject_map.get(test_type, test_type)

        # Store/overwrite subject score (only one test per subject per week, latest wins)
        student_data[student_id]["scores"][subject_name] = {
            "subject": subject_name,
            "mark": test_result.total_score,
            "max_mark": test_result.max_score,
            "percentage": round(test_result.percentage, 2) if test_result.percentage else None,
            "test_id": str(test.id),
            "test_title": test.title,
            "submitted_at": test_result.submitted_at.isoformat() if test_result.submitted_at else None
        }

    # Build response
    students_list = []
    for student_id, data in sorted(
        student_data.items(),
        key=lambda x: (x[1]["info"].get("class_name", ""), x[1]["info"].get("surname", ""))
    ):
        info = data["info"]
        scores = data["scores"]

        # Ensure all subjects are present (even if no score)
        all_scores = {}
        for subject in subjects:
            if subject in scores:
                all_scores[subject] = SubjectScore(**scores[subject])
            else:
                all_scores[subject] = SubjectScore(
                    subject=subject,
                    mark=None,
                    max_mark=None,
                    percentage=None,
                    test_id=None,
                    test_title=None,
                    submitted_at=None
                )

        students_list.append(
            StudentWeeklyScores(
                student_id=info["student_id"],
                student_code=info["student_code"],
                first_name=info["first_name"],
                surname=info["surname"],
                full_name=info["full_name"],
                class_id=info["class_id"],
                class_name=info["class_name"],
                year_group=info["year_group"],
                scores=all_scores
            )
        )

    # Create week info response
    week_info_response = AcademicWeekInfo(
        week_number=week_info.week_number,
        start_date=week_info.start_date.isoformat(),
        end_date=week_info.end_date.isoformat(),
        is_break=week_info.is_break,
        break_name=week_info.break_name,
        week_label=calendar_service.get_week_label(target_week)
    )

    result = WeeklyResultsSummary(
        week_info=week_info_response,
        students=students_list,
        total_students=len(students_list),
        subjects=subjects
    )

    # Return with camelCase aliases for frontend
    return result.model_dump(by_alias=True)


@router.get("/results/export")
async def export_results(
    db: AsyncSession = Depends(get_db),
    current_admin: User = Depends(get_current_admin_user)
):
    """
    Export all test results in 40-week format matching Excel template.

    Format:
    - Row 1: Week headers (Week1, Week2, ..., Week40)
    - Row 2: Subject names under each week (English, VR GL, NVR, Maths)
    - Row 3: Column headers (Class ID, Student ID, First Name, Surname, then Mark/% for each subject per week)
    - Data rows: Student information + scores organized by week and subject
    """
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import datetime
    from sqlalchemy import select
    from app.models.test import TestResult, Test
    from app.models.student import Student
    from sqlalchemy.orm import selectinload
    from app.models.user import User
    from app.models.class_model import Class
    from app.services.academic_calendar_service import calendar_service
    from collections import defaultdict

    # Fetch all results with student and test details
    query = (
        select(TestResult)
        .options(
            selectinload(TestResult.student).selectinload(Student.user),
            selectinload(TestResult.student).selectinload(Student.class_info),
            selectinload(TestResult.test)
        )
        .order_by(TestResult.submitted_at)
    )

    result = await db.execute(query)
    all_results = result.scalars().all()

    # Organize results by student and week
    # Structure: {student_id: {week_num: {subject: {mark, percentage, ...}}}}
    student_data = defaultdict(lambda: {"info": {}, "weeks": defaultdict(lambda: defaultdict(dict))})

    for test_result in all_results:
        if not test_result.student or not test_result.test or not test_result.submitted_at:
            continue

        student = test_result.student
        test = test_result.test

        # Determine academic week from submission date
        submission_date = test_result.submitted_at.date()
        week_number = calendar_service.date_to_week_number(submission_date)

        # Skip if outside academic calendar
        if week_number == 0:
            continue

        # Extract student info (only once per student)
        student_id = str(student.id)
        if not student_data[student_id]["info"]:
            # Split full name into first name and surname
            full_name = student.user.full_name if student.user else ""
            name_parts = full_name.strip().split(maxsplit=1)
            first_name = name_parts[0] if len(name_parts) > 0 else ""
            surname = name_parts[1] if len(name_parts) > 1 else ""

            student_data[student_id]["info"] = {
                "class_id": student.class_info.name if student.class_info else "N/A",
                "student_code": student.student_code or "N/A",
                "first_name": first_name,
                "surname": surname,
                "year_group": student.year_group
            }

        # Map test type to subject name (matching Excel template)
        test_type = test.type.value if hasattr(test.type, 'value') else str(test.type)
        subject_map = {
            "English": "English",
            "Verbal Reasoning": "VR GL",
            "Non-Verbal Reasoning": "NVR",
            "Mathematics": "Maths"
        }
        subject_name = subject_map.get(test_type, test_type)

        # Store/overwrite result for this week and subject (only one test per subject per week, latest wins)
        student_data[student_id]["weeks"][week_number][subject_name] = {
            "mark": test_result.total_score or 0,
            "percentage": round(test_result.percentage, 2) if test_result.percentage else 0
        }

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Results"

    # Define styles
    header_fill = PatternFill(start_color="DB2E1D", end_color="DB2E1D", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    # Thick border for week end separator (only right border is thick)
    week_end_border = Border(
        # left=Side(style='thin'),
        right=Side(style='thick', color='000000'),  # Thick black border on the right only
        top=Side(style='thin'),
        # bottom=Side(style='thin')
    )

    # Subjects in order (matching Excel template)
    subjects = ["English", "VR GL", "NVR", "Maths"]

    # Calculate total columns: 4 (student info) + 40 weeks × 4 subjects × 2 columns (Mark + %)
    # = 4 + 320 = 324 columns total

    current_col = 1

    # Row 1: Week headers
    # Columns A-D: Empty for student info
    current_col = 5  # Start after student info columns (A-D)

    for week_num in range(1, 41):
        # Each week spans 8 columns (4 subjects × 2 columns each)
        week_label = f"Week{week_num}"
        week_end_col = current_col + 7

        # Set week header in first cell
        cell = ws.cell(row=1, column=current_col)
        cell.value = week_label
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

        # Merge cells for week header (8 columns: 4 subjects × 2 fields)
        ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=week_end_col)

        # Apply thick border to last cell of each week header
        for col in range(current_col, week_end_col + 1):
            header_cell = ws.cell(row=1, column=col)
            if col == week_end_col:
                header_cell.border = week_end_border
            else:
                header_cell.border = border

        current_col += 8

    # Row 2: Subject names
    current_col = 5
    for week_num in range(1, 41):
        for subject_idx, subject in enumerate(subjects):
            is_last_subject = (subject_idx == len(subjects) - 1)
            subject_end_col = current_col + 1

            cell = ws.cell(row=2, column=current_col)
            cell.value = subject
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

            # Merge 2 columns for subject name (Mark and %)
            ws.merge_cells(start_row=2, start_column=current_col, end_row=2, end_column=subject_end_col)

            # Apply thick border to last subject column of each week
            for col in range(current_col, subject_end_col + 1):
                subject_cell = ws.cell(row=2, column=col)
                if is_last_subject and col == subject_end_col:
                    subject_cell.border = week_end_border
                else:
                    subject_cell.border = border

            current_col += 2

    # Row 3: Column headers (Mark and %)
    # Student info headers
    student_headers = ["Class ID", "Student ID", "First Name", "Surname"]
    for col_num, header in enumerate(student_headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # Mark and % headers for each subject in each week
    current_col = 5
    for week_num in range(1, 41):
        for subject_idx, subject in enumerate(subjects):
            is_last_subject = (subject_idx == len(subjects) - 1)

            # Mark column
            cell = ws.cell(row=3, column=current_col)
            cell.value = "Mark"
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

            # % column (last column of week gets thick border)
            cell = ws.cell(row=3, column=current_col + 1)
            cell.value = "%"
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            # Apply thick border to last column of each week
            if is_last_subject:
                cell.border = week_end_border
            else:
                cell.border = border

            current_col += 2

    # Data rows (starting from row 4)
    row_num = 4
    for student_id, data in sorted(student_data.items(), key=lambda x: (x[1]["info"].get("class_id", ""), x[1]["info"].get("surname", ""))):
        info = data["info"]
        weeks = data["weeks"]

        # Student info columns
        ws.cell(row=row_num, column=1, value=info["class_id"]).alignment = data_alignment
        ws.cell(row=row_num, column=2, value=info["student_code"]).alignment = data_alignment
        ws.cell(row=row_num, column=3, value=info["first_name"]).alignment = data_alignment
        ws.cell(row=row_num, column=4, value=info["surname"]).alignment = data_alignment

        # Scores for each week and subject
        current_col = 5
        for week_num in range(1, 41):
            week_data = weeks.get(week_num, {})
            for subject_idx, subject in enumerate(subjects):
                is_last_subject = (subject_idx == len(subjects) - 1)
                subject_data = week_data.get(subject, {})

                # Mark
                mark = subject_data.get("mark", "")
                mark_cell = ws.cell(row=row_num, column=current_col, value=mark if mark != "" else "")
                mark_cell.alignment = data_alignment

                # Percentage (last column of week gets thick border)
                percentage = subject_data.get("percentage", "")
                pct_cell = ws.cell(row=row_num, column=current_col + 1, value=percentage if percentage != "" else "")
                pct_cell.alignment = data_alignment
                # Apply thick border to last column of each week
                if is_last_subject:
                    pct_cell.border = week_end_border

                current_col += 2

        row_num += 1

    # Set column widths
    # Student info columns
    ws.column_dimensions['A'].width = 12  # Class ID
    ws.column_dimensions['B'].width = 15  # Student ID
    ws.column_dimensions['C'].width = 15  # First Name
    ws.column_dimensions['D'].width = 15  # Surname

    # Week columns (Mark and % columns)
    from openpyxl.utils import get_column_letter
    for col_num in range(5, 5 + (40 * 8)):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 8

    # Freeze top 3 rows and first 4 columns
    ws.freeze_panes = 'E4'

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Generate filename with current date and timestamp
    filename = f"test-results-{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


# Academic Calendar Configuration Endpoints

@router.get("/academic-calendar/current-week")
async def get_current_academic_week(
    current_admin: User = Depends(get_current_admin_user)
):
    """Get information about the current academic week."""
    from app.services.academic_calendar_service import calendar_service
    from app.schemas.analytics import AcademicWeekInfo

    current_week_num = calendar_service.get_current_week()

    if current_week_num == 0:
        return {
            "current_week": 0,
            "message": "No active academic week",
            "academic_year": calendar_service.get_academic_year_string()
        }

    week_info = calendar_service.get_week_info(current_week_num)

    week_info_obj = AcademicWeekInfo(
        week_number=week_info.week_number,
        start_date=week_info.start_date.isoformat(),
        end_date=week_info.end_date.isoformat(),
        is_break=week_info.is_break,
        break_name=week_info.break_name,
        week_label=calendar_service.get_week_label(current_week_num)
    )

    return {
        "currentWeek": current_week_num,
        "weekInfo": week_info_obj.model_dump(by_alias=True),
        "academicYear": calendar_service.get_academic_year_string(),
        "totalWeeks": calendar_service.TOTAL_WEEKS
    }


@router.get("/academic-calendar/all-weeks")
async def get_all_academic_weeks(
    current_admin: User = Depends(get_current_admin_user)
):
    """Get information about all 40 academic weeks."""
    from app.services.academic_calendar_service import calendar_service
    from app.schemas.analytics import AcademicWeekInfo

    all_weeks_info = calendar_service.get_all_weeks_info()
    current_week = calendar_service.get_current_week()

    weeks_response = []
    for week_info in all_weeks_info:
        weeks_response.append(AcademicWeekInfo(
            week_number=week_info.week_number,
            start_date=week_info.start_date.isoformat(),
            end_date=week_info.end_date.isoformat(),
            is_break=week_info.is_break,
            break_name=week_info.break_name,
            week_label=calendar_service.get_week_label(week_info.week_number)
        ))

    return {
        "academic_year": calendar_service.get_academic_year_string(),
        "current_week": current_week,
        "total_weeks": calendar_service.TOTAL_WEEKS,
        "weeks": weeks_response
    }


@router.get("/academic-calendar/config")
async def get_academic_calendar_config(
    db: AsyncSession = Depends(get_db),
    current_admin: User = Depends(get_current_admin_user)
):
    """Get the current academic calendar configuration."""
    from app.models.academic_calendar import AcademicCalendarConfig
    from sqlalchemy import select
    from app.services.academic_calendar_service import calendar_service

    # Try to get config from database
    query = select(AcademicCalendarConfig).where(
        AcademicCalendarConfig.is_active == True
    ).order_by(AcademicCalendarConfig.created_at.desc())

    result = await db.execute(query)
    config = result.scalar_one_or_none()

    if not config:
        # Return default configuration
        return {
            "academic_year": calendar_service.get_academic_year_string(),
            "year_start_date": calendar_service.ACADEMIC_YEAR_START.isoformat(),
            "total_weeks": calendar_service.TOTAL_WEEKS,
            "break_periods": [
                {
                    "name": bp.name,
                    "start_date": bp.start_date.isoformat(),
                    "end_date": bp.end_date.isoformat()
                }
                for bp in calendar_service.DEFAULT_BREAKS
            ],
            "week_start_day": "Friday",
            "week_end_day": "Wednesday",
            "is_active": True,
            "notes": "Default configuration"
        }

    # Return database configuration
    from app.schemas.analytics import AcademicCalendarConfigSchema
    return AcademicCalendarConfigSchema.model_validate(config)


@router.put("/academic-calendar/config")
async def update_academic_calendar_config(
    config_data: dict,
    db: AsyncSession = Depends(get_db),
    current_admin: User = Depends(get_current_admin_user)
):
    """
    Update the academic calendar configuration (primarily break periods).

    Note: This is simplified for MVP. For production, use proper schema validation.
    """
    from app.models.academic_calendar import AcademicCalendarConfig
    from sqlalchemy import select

    # Get or create config
    query = select(AcademicCalendarConfig).where(
        AcademicCalendarConfig.is_active == True
    ).order_by(AcademicCalendarConfig.created_at.desc())

    result = await db.execute(query)
    config = result.scalar_one_or_none()

    if not config:
        # Create new config
        from app.services.academic_calendar_service import calendar_service

        config = AcademicCalendarConfig(
            academic_year=calendar_service.get_academic_year_string(),
            year_start_date=calendar_service.ACADEMIC_YEAR_START,
            total_weeks=calendar_service.TOTAL_WEEKS,
            break_periods=config_data.get("break_periods", []),
            notes=config_data.get("notes"),
            created_by=current_admin.id
        )
        db.add(config)
    else:
        # Update existing config
        if "break_periods" in config_data:
            config.break_periods = config_data["break_periods"]
        if "notes" in config_data:
            config.notes = config_data["notes"]

    await db.commit()
    await db.refresh(config)

    return {"message": "Configuration updated successfully", "config_id": str(config.id)}